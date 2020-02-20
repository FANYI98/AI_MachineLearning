#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Nov 30 01:35:09 2019

@author: yifan
"""
from openpyxl import *
wb=load_workbook(filename='./raw_data.xlsx')
sheet=wb['raw_data']

import sys
AIMA_TOOLBOX_ROOT="./aima-python-uofg_v20192020b"
sys.path.append(AIMA_TOOLBOX_ROOT)

from helpers import *
from search import *
from Astar_search import *
from uofgsocsai import LochLomondEnv 

problem_id = int(sys.argv[1])
map_name_base = sys.argv[2]

#initialise the environment
env = LochLomondEnv(problem_id=problem_id, is_stochastic=False, map_name_base=map_name_base, reward_hole=0.0)

# visualize the problem/env
print(env.desc)

#define the frozen_lake Graph problem
state_space_locations, state_space_actions, state_initial_id, state_goal_id = env2statespace(env)
frozen_lake_map = UndirectedGraph(state_space_actions)
frozen_lake_map.locations = state_space_locations
frozen_lake_problem = GraphProblem(state_initial_id, state_goal_id, frozen_lake_map)

#find the solution
all_node_colors=[]
iterations, all_node_colors, node = my_astar_search_graph(problem=frozen_lake_problem, h=None)
print("Number of steps: "+str(iterations))

#Trace the solution 
solution_path = [node]
cnode = node.parent
solution_path.append(cnode)
while cnode.state != "S_00_00":    
    cnode = cnode.parent
    if cnode is None:
        break
    solution_path.append(cnode)

print("----------------------------------------")
print("Identified goal state:"+str(solution_path[0]))
print("----------------------------------------")
print("Solution trace:"+str(solution_path))
print("----------------------------------------")

#write the evaluation parameters into raw_data xlsx
if map_name_base=='8x8-base':
    sheet.cell(3,3*(problem_id+1)+1).value = 1
    sheet.cell(4,3*(problem_id+1)+1).value = 1
    sheet.cell(5,3*(problem_id+1)+1).value = 1
    sheet.cell(6,3*(problem_id+1)+1).value = iterations
    sheet.cell(7,3*(problem_id+1)+1).value = iterations
if map_name_base=='4x4-base':
    sheet.cell(8,3*(problem_id+1)+1).value = 1
    sheet.cell(9,3*(problem_id+1)+1).value = 1
    sheet.cell(10,3*(problem_id+1)+1).value = 1
    sheet.cell(11,3*(problem_id+1)+1).value = iterations
    sheet.cell(12,3*(problem_id+1)+1).value = iterations
wb.save('./raw_data.xlsx')

print('Relevant parameters has been written to raw_data.xlsx!')