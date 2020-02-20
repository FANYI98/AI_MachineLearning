#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Nov 30 01:42:38 2019

@author: yifan
"""
from openpyxl import *
wb=load_workbook(filename='./raw_data.xlsx')
sheet=wb['raw_data']

import sys
from uofgsocsai import LochLomondEnv
AIMA_TOOLBOX_ROOT="./aima-python-uofg_v20192020b"
sys.path.append(AIMA_TOOLBOX_ROOT)

import numpy as np
from search import *
import matplotlib.pyplot as plt

#change the action of each state according to Qtable into arrow
def switch_arrow(action):
        switcher = {
            0: "v",
            1: "<",
            2: "^",
            3: ">",    
        }
        return switcher.get(action,".")
    
problem_id = int(sys.argv[1])
map_name_base = sys.argv[2]
#problem_id =1
#map_name_base = "4x4-base" 
max_episodes=10000
max_iter_per_episode=300

lr = .8 #learning rate
y = .95 #discounting rate

ave_steps_success=0 #average steps among all successful episodes
rewardAll=0 #record the reward of each episodes
rewardL=[] #list of rewards of each episodes
k=0 #record the number of successful episodes
step_number = [] #record the steps of each successful episode

#initialize the environment
env = LochLomondEnv(problem_id=problem_id, is_stochastic=True, map_name_base=map_name_base, reward_hole=-0.4)

#initialize Q tabble with states many rows and actions many columns 
Q = np.zeros([env.observation_space.n,env.action_space.n])

for i in range(max_episodes):
    #Reset environment and get first new observation   
    state = env.reset()
    rewardAll= 0
    done = False
    #The Q-Table learning algorithm
    for j in range(max_iter_per_episode):
        #Choose an action a in the current world state (s) with Greedy             
        action = np.argmax(Q[state,:] + np.random.randn(1, env.action_space.n) * (1./(i+1)))        
        #Get new state and reward from environment
        state1,reward,done,_ = env.step(action)
        #Update Q-Table with new knowledge
        Q[state,action] = Q[state,action] + lr*(reward + y*np.max(Q[state1,:]) - Q[state,action])
        rewardAll += reward
        state = state1
        if done == True:
            if (rewardAll == 1):
                k=k+1
                step_number.append(j)                       
            break
    rewardL.append(rewardAll)

print("Success rates: " +  str(k/max_episodes))
print( "Score over time: " +  str(sum(rewardL)/max_episodes))
print("Final Q-Table Values")
print(Q)

#get the action of each state according to Q table and show in arrow
if (map_name_base == "8x8-base" ):
    unit = 8
if (map_name_base == "4x4-base" ):
    unit = 4
arr = np.empty((unit,unit), dtype=object)
for s in range (env.observation_space.n):
       row=int(s/unit)
       colum=int(s%unit)
       action = np.argmax(Q[s,:])
       if (Q[s,0]==Q[s,1]==Q[s,2]==Q[s,3]==0):
           action = None
       arr[row,colum] = switch_arrow(action)  
print('single policy according to Q table : ')     
print(arr)

#record the average reward per 100 episodes   
ave_reward=np.zeros(int(len(rewardL)/100))
for i in range(0,int(len(rewardL)/100)):   
    ave_reward[i]=sum(rewardL[100*i:100*(i+1)])/100
plt.plot(ave_reward)
plt.title("average reward(/100 episodes) of RL agent")
plt.show()
   
if len(step_number)!=0:
    ave_steps_success=sum(step_number)/len(step_number)

#write the evaluation parameters into raw_data xlsx
if map_name_base=='8x8-base':
    sheet.cell(3,3*(problem_id+1)+2).value = ",".join(str(e)for e in ave_reward)
    sheet.cell(4,3*(problem_id+1)+2).value = sum(rewardL)/max_episodes
    sheet.cell(5,3*(problem_id+1)+2).value = str(k/max_episodes)
    sheet.cell(6,3*(problem_id+1)+2).value = ",".join(str(e)for e in step_number)
    sheet.cell(7,3*(problem_id+1)+2).value = ave_steps_success
if map_name_base=='4x4-base':
    sheet.cell(8,3*(problem_id+1)+2).value = ",".join(str(e)for e in ave_reward)
    sheet.cell(9,3*(problem_id+1)+2).value = sum(rewardL)/max_episodes
    sheet.cell(10,3*(problem_id+1)+2).value = str(k/max_episodes)
    sheet.cell(11,3*(problem_id+1)+2).value = ",".join(str(e)for e in step_number)
    sheet.cell(12,3*(problem_id+1)+2).value = ave_steps_success
wb.save('./raw_data.xlsx')

print('Relevant parameters has been written to raw_data.xlsx!')
