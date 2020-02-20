#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 28 15:51:32 2019

@author: yifan
"""

import numpy as np
import sys
import matplotlib.pyplot as plt
from openpyxl import *
#map_name_base='8x8-base'
map_name_base = sys.argv[1]
wb1 = load_workbook(filename='./raw_data.xlsx')
sheet1 = wb1['raw_data']

#initialize a new excel to summarise the results over the 8 instances of the problem 
wb2 = Workbook()
sheet2 = wb2.active
sheet2.column_dimensions['B'].width = 30
sheet2.title = 'results_summarising'

#copy the relevant raw_data into the results_summarising excel
max_column=sheet1.max_column
if map_name_base=='8x8-base':
    for m in range(1,8):
        for n in range(97,97+max_column):#chr(97)='a'
            n=chr(n)#ASCII
            i='%s%d'%(n,m)#cell number
            cell1=sheet1[i].value#get the data of cell
            sheet2[i].value=cell1#set the data into cell in new xlsx
if map_name_base=='4x4-base':    
    for m in range(1,3):
        for n in range(97,97+max_column):
            n=chr(n)
            i='%s%d'%(n,m)
            cell1=sheet1[i].value
            sheet2[i].value=cell1
    for m in range(8,13):
        for n in range(97,97+max_column):
            n=chr(n)
            i='%s%d'%(n,m)
            j='%s%d'%(n,m-5)
            cell1=sheet1[i].value
            sheet2[j].value=cell1
wb2.save('./results_summarising.xlsx')
wb1.close()
wb2.close()
print('=====================================')
print('A new excel named(results_summarising.xlsx)summarising the results has been created!')

#plot learning behavior and convergence of the 3 agents among 8 performs
if map_name_base=='8x8-base':
    unit=8
if map_name_base=='4x4-base':
    unit=4
#read the ave_reward value from the results_summarising.xlsx into array
#comparison among the 3 agents in terms of ave_reward of each problem
array_random=np.zeros([unit,100])
array_simple=np.zeros([unit,100])
array_rl=np.zeros([unit,100])
for i in range (0,unit):
    str_value=sheet2.cell(row=3,column=3*(i+1)).value    
    array_random[i,:]=eval(str_value)
    str_value1=sheet2.cell(row=3,column=3*(i+1)+1).value    
    array_simple[i]=str_value1
    str_value2=sheet2.cell(row=3,column=3*(i+1)+2).value    
    array_rl[i,:]=eval(str_value2)
    plt.figure(i)
    plt.plot(array_random[i,:],label='random_agent')
    plt.plot(array_simple[i],label='simple_agent')
    plt.plot(array_rl[i,:],label='RL_agent')
    plt.legend()
    plt.title('problem'+str(i))
    plt.ylabel('ave_reward/100 episodes')
    plt.xlabel('/100 episodes')
    plt.show()
    
#plot the random agent ave_reward of 8 problems  
plt.figure(8)
for i in range (0,unit):
    plt.plot(array_random[i,:],label='problem'+str(i))
plt.legend()
plt.title('random_agent')
plt.ylabel('ave_reward/100 episodes')
plt.xlabel('/100 episodes')
plt.show()
#plot the rl agent ave_reward of 8 problems 
plt.figure(9)
for i in range (0,unit):
    plt.plot(array_rl[i,:],label='problem'+str(i))
plt.legend()
plt.title('rl_agent')
plt.ylabel('ave_reward/100 episodes')
plt.xlabel('/100 episodes')
plt.show()
#plot the simple agent ave_reward of 8 problems 
plt.figure(10)
for i in range (0,unit):
    plt.plot(array_simple[i,:],label='problem'+str(i))
plt.legend()
plt.title('simple_agent')
plt.ylabel('ave_reward/100 episodes')
plt.xlabel('/100 episodes')
plt.show()


