#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Nov 30 01:21:26 2019
@author: yifan
"""
from openpyxl import *
wb=load_workbook(filename='./raw_data.xlsx')
sheet=wb['raw_data']
import sys
import numpy as np
from uofgsocsai import LochLomondEnv 
import matplotlib.pyplot as plt

problem_id = int(sys.argv[1])
map_name_base = sys.argv[2]
max_episodes=10000
max_iter_per_episode =300

k=0 #record the number of successful episodes

rewardL=[]# record the reward of each episode
ave_reward_total=0 #average reward among max_episodes
success_steps=[] #record the steps of each successful episode
ave_steps_success=0 #average steps among all successful episodes
#initialize the environment
env = LochLomondEnv(problem_id=problem_id, is_stochastic=True, map_name_base=map_name_base, reward_hole=0.0)

#random/senseless agent
for e in range(max_episodes):
    observation = env.reset() 
    for iter in range(max_iter_per_episode):
      action = env.action_space.sample() 
      observation, reward, done, info = env.step(action)
      if(done):
          rewardL.append(reward)
          if(reward==1):
              k=k+1
              success_steps.append(iter)
              env.render()     
              print("episodes,iter,done =" + " "+ str(e) + " " + str(iter)+ " " + str(done))
              print("We arrive at the gate. Please try again!")
          break

print("success rate = " + str(k/max_episodes))

#record the average reward per 100 episodes
ave_reward=np.zeros(int(len(rewardL)/100))
for i in range(0,int(len(rewardL)/100)):
    ave_reward[i]=sum(rewardL[i*100:(i+1)*100])/100
plt.plot(ave_reward)
plt.title("average reward(/100 episodes) of random agent")
plt.show()

ave_reward_total=sum(ave_reward)/len(ave_reward)
if len(success_steps)!=0:
    ave_steps_success=sum(success_steps)/len(success_steps)

#write the evaluation parameters into raw_data xlsx
if map_name_base=='8x8-base':
    sheet.cell(3,3*(problem_id+1)).value = ",".join(str(e)for e in ave_reward)
    sheet.cell(4,3*(problem_id+1)).value = ave_reward_total
    sheet.cell(5,3*(problem_id+1)).value = str(k/max_episodes)
    sheet.cell(6,3*(problem_id+1)).value = ",".join(str(e)for e in success_steps)
    sheet.cell(7,3*(problem_id+1)).value = ave_steps_success
if map_name_base=='4x4-base':
    sheet.cell(8,3*(problem_id+1)).value = ",".join(str(e)for e in ave_reward)
    sheet.cell(9,3*(problem_id+1)).value = ave_reward_total
    sheet.cell(10,3*(problem_id+1)).value = str(k/max_episodes)
    sheet.cell(11,3*(problem_id+1)).value = ",".join(str(e)for e in success_steps)
    sheet.cell(12,3*(problem_id+1)).value =ave_steps_success
wb.save('./raw_data.xlsx')
print('Relevant parameters has been written to raw_data.xlsx!')