#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Nov 30 03:30:27 2019

@author: yifan
"""
#initialize the excel table used for save relevant data
from openpyxl import *
from openpyxl.styles import Alignment
excel_raw = Workbook()
sheet = excel_raw.active
sheet.title = 'raw_data'
sheet.column_dimensions['B'].width = 30
sheet.merge_cells('C1:E1')
sheet.merge_cells('F1:H1')
sheet.merge_cells('I1:K1')
sheet.merge_cells('L1:N1')
sheet.merge_cells('O1:Q1')
sheet.merge_cells('R1:T1')
sheet.merge_cells('U1:W1')
sheet.merge_cells('X1:Z1')

for i in range (8):
    sheet.cell(1,(i+1)*3,"problem"+str(i))
    sheet.cell(2,(i+1)*3,"random_agent")
    sheet.cell(2,(i+1)*3+1,"simple_agent")
    sheet.cell(2,(i+1)*3+2,"RL_agent")

sheet.merge_cells('A3:A7')
sheet.cell(3,1).value = '8x8-base'
sheet.cell(3,2,"ave_reward/100 episodes")
sheet.cell(4,2,"ave_reward_total")
sheet.cell(5,2,"success_rates")
sheet.cell(6,2,"items for success")
sheet.cell(7,2,"ave_stps_success")


sheet.merge_cells('A8:A12')
sheet.cell(8,1).value = '4x4-base'
sheet.cell(8,2,"ave_reward/100 episodes")
sheet.cell(9,2,"ave_reward_total")
sheet.cell(10,2,"success_rates")
sheet.cell(11,2,"items for success")
sheet.cell(12,2,"ave_stps_success")

path = "./raw_data.xlsx"
excel_raw.save(path)